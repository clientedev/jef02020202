from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from sqlalchemy.orm import Session
from typing import List, Optional
from backend.database import get_db
from backend.models import Empresa, Usuario
from backend.schemas.empresas import EmpresaCriar, EmpresaResposta, EmpresaAtualizar
from backend.auth.security import obter_usuario_atual, obter_usuario_admin
import openpyxl
from io import BytesIO
from datetime import datetime

router = APIRouter(prefix="/api/empresas", tags=["Empresas"])

# Import lookup logic from cnpj.py
from backend.routers.cnpj import buscar_empresa_cnpj, limpar_cnpj as sanitizar_cnpj

@router.post("/", response_model=EmpresaResposta)
def criar_empresa(
    empresa: EmpresaCriar,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_admin)
):
    if empresa.cnpj:
        db_empresa = db.query(Empresa).filter(Empresa.cnpj == empresa.cnpj).first()
        if db_empresa:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Empresa com este CNPJ já cadastrada"
            )
    
    nova_empresa = Empresa(**empresa.model_dump())
    db.add(nova_empresa)
    db.commit()
    db.refresh(nova_empresa)
    return nova_empresa

@router.get("/")
def listar_empresas(
    page: int = Query(1, ge=1, description="Número da página"),
    page_size: int = Query(20, ge=1, le=100, description="Items por página"),
    nome: Optional[str] = None,
    cnpj: Optional[str] = None,
    municipio: Optional[str] = None,
    er: Optional[str] = None,
    carteira: Optional[str] = None,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual)
):
    query = db.query(Empresa)
    
    if nome:
        query = query.filter(Empresa.empresa.ilike(f"%{nome}%"))
    if cnpj:
        query = query.filter(Empresa.cnpj.ilike(f"%{cnpj}%"))
    if municipio:
        query = query.filter(Empresa.municipio.ilike(f"%{municipio}%"))
    if er:
        query = query.filter(Empresa.er == er)
    if carteira:
        query = query.filter(Empresa.carteira == carteira)
    
    total_count = query.count()
    total_pages = (total_count + page_size - 1) // page_size
    
    skip = (page - 1) * page_size
    empresas = query.offset(skip).limit(page_size).all()
    
    from backend.schemas.empresas import EmpresaResposta
    items_safe = [EmpresaResposta.model_validate(e) for e in empresas]
    
    return {
        "items": items_safe,
        "total_count": total_count,
        "page": page,
        "page_size": page_size,
        "total_pages": total_pages
    }

@router.get("/{empresa_id}", response_model=EmpresaResposta)
def obter_empresa(
    empresa_id: int,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual)
):
    empresa = db.query(Empresa).filter(Empresa.id == empresa_id).first()
    if not empresa:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empresa não encontrada"
        )
    return empresa

@router.get("/{empresa_id}/ultimo-contato")
def obter_ultimo_contato(
    empresa_id: int,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual)
):
    from backend.models.prospeccoes import Prospeccao
    
    empresa = db.query(Empresa).filter(Empresa.id == empresa_id).first()
    if not empresa:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empresa não encontrada"
        )
    
    nome = empresa.nome_contato
    cargo = empresa.cargo_contato
    telefone = empresa.telefone_contato
    email = empresa.email_contato
    
    if not any([nome, cargo, telefone, email]):
        ultima_prospeccao = db.query(Prospeccao).filter(
            Prospeccao.empresa_id == empresa_id
        ).order_by(Prospeccao.data_criacao.desc()).first()
        
        if ultima_prospeccao:
            nome = nome or ultima_prospeccao.nome_contato
            cargo = cargo or ultima_prospeccao.cargo_contato or ultima_prospeccao.cargo
            telefone = telefone or ultima_prospeccao.telefone_contato or ultima_prospeccao.telefone or ultima_prospeccao.celular
            email = email or ultima_prospeccao.email_contato
    
    tem_contato = bool(nome or cargo or telefone or email)
    
    return {
        "tem_contato": tem_contato,
        "nome_contato": nome,
        "cargo_contato": cargo,
        "telefone_contato": telefone,
        "email_contato": email
    }

@router.put("/{empresa_id}", response_model=EmpresaResposta)
def atualizar_empresa(
    empresa_id: int,
    empresa_atualizada: EmpresaAtualizar,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_admin)
):
    empresa = db.query(Empresa).filter(Empresa.id == empresa_id).first()
    if not empresa:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empresa não encontrada"
        )
    
    for key, value in empresa_atualizada.model_dump(exclude_unset=True).items():
        setattr(empresa, key, value)
    
    db.commit()
    db.refresh(empresa)
    return empresa

@router.delete("/limpar-tudo")
def limpar_todas_empresas(
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_admin)
):
    from backend.models.prospeccoes import Prospeccao, ProspeccaoHistorico
    from backend.models.agendamentos import Agendamento
    from backend.models.atribuicoes import AtribuicaoEmpresa
    from backend.models.cronograma import CronogramaProjeto, CronogramaAtividade, CronogramaEvento
    from backend.models.pipeline import CompanyPipeline, CompanyStageHistory, Note, Attachment, Activity
    from backend.models.historico import HistoricoEmpresa
    from backend.models.contatos import Contato

    # Limpar dependências que não têm cascade automático ou precisam de limpeza manual
    db.query(Activity).delete()
    db.query(Attachment).delete()
    db.query(Note).delete()
    db.query(CompanyStageHistory).delete()
    db.query(CompanyPipeline).delete()
    db.query(CronogramaEvento).delete()
    db.query(CronogramaAtividade).delete()
    db.query(CronogramaProjeto).delete()
    db.query(AtribuicaoEmpresa).delete()
    db.query(Agendamento).delete()
    db.query(ProspeccaoHistorico).delete()
    db.query(Prospeccao).delete()
    
    # As empresas e seus contatos/historico (que têm cascade no modelo)
    db.query(Empresa).delete()
    
    db.commit()
    return {"message": "Todas as empresas e dados relacionados foram removidos com sucesso"}

@router.delete("/{empresa_id}")
def deletar_empresa(
    empresa_id: int,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_admin)
):
    empresa = db.query(Empresa).filter(Empresa.id == empresa_id).first()
    if not empresa:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Empresa não encontrada"
        )
    
    db.delete(empresa)
    db.commit()
    return {"detail": "Empresa deletada com sucesso"}

@router.post("/sync-cnpj")
async def sincronizar_cnpj_empresas(
    empresa_id: Optional[int] = None,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_admin)
):
    """
    Sincroniza dados via CNPJ para uma empresa específica ou para todas as incompletas.
    """
    if empresa_id:
        empresas_para_sync = db.query(Empresa).filter(Empresa.id == empresa_id).all()
    else:
        # Busca empresas que tem CNPJ mas faltam dados básicos (municipio ou endereco)
        empresas_para_sync = db.query(Empresa).filter(
            Empresa.cnpj.isnot(None),
            (Empresa.municipio.is_(None)) | (Empresa.endereco.is_(None))
        ).limit(20).all() # Limite de 20 por vez para não estourar APIs

    if not empresas_para_sync:
        return {"message": "Nenhuma empresa pendente de sincronização encontrada.", "sincronizados": 0}

    sincronizados = 0
    erros = 0

    for emp in empresas_para_sync:
        if not emp.cnpj:
            erros += 1
            continue
            
        try:
            # Chama a lógica de busca que já alterna entre ReceitaWS, BrasilAPI e CNPJA
            dados = await buscar_empresa_cnpj(emp.cnpj, usuario)
            
            # Atualiza os campos apenas se estiverem vazios ou para enriquecer o cadastro
            emp.empresa = emp.empresa or dados.get("empresa")
            emp.sigla = emp.sigla or dados.get("nome_fantasia")
            emp.municipio = emp.municipio or dados.get("municipio")
            emp.estado = emp.estado or dados.get("estado")
            emp.bairro = emp.bairro or dados.get("bairro")
            emp.cep = emp.cep or dados.get("cep")
            emp.porte = emp.porte or dados.get("porte")
            emp.tipo_empresa = emp.tipo_empresa or dados.get("natureza_juridica")
            emp.descricao_cnae = emp.descricao_cnae or dados.get("atividade_principal")
            
            # Formata endereço completo se necessário
            if not emp.endereco and dados.get("logradouro"):
                addr = dados.get("logradouro")
                if dados.get("numero"): addr += f", {dados.get('numero')}"
                if dados.get("complemento"): addr += f" - {dados.get('complemento')}"
                emp.endereco = addr

            sincronizados += 1
        except Exception as e:
            print(f"Erro sincronizando CNPJ {emp.cnpj}: {e}")
            erros += 1

    db.commit()
    return {
        "message": f"Sincronização concluída: {sincronizados} atualizadas, {erros} falhas.",
        "sincronizados": sincronizados,
        "falhas": erros
    }

@router.post("/upload-excel")
async def upload_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_admin)
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Arquivo deve ser Excel (.xlsx ou .xls)"
        )
    
    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(BytesIO(contents))
        ws = wb.active
        
        headers = [str(cell.value).strip().upper() if cell in ws[1] and cell.value else "" for cell in ws[1]]
        
        def get_col_val(row, *possible_names):
            for name in possible_names:
                name_upper = name.upper()
                if name_upper in headers:
                    idx = headers.index(name_upper)
                    val = row[idx]
                    return str(val).strip() if val is not None and str(val).strip().lower() != 'nan' else None
            return None

        empresas_criadas = 0
        empresas_ignoradas = 0
        cnpjs_processados = set()
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            empresa_nome = get_col_val(row, 'EMPRESA', 'NOME DA EMPRESA', 'NOME')
            cnpj = get_col_val(row, 'CNPJ')
            
            if not empresa_nome:
                continue
            
            if cnpj and cnpj in cnpjs_processados:
                empresas_ignoradas += 1
                continue
                
            if cnpj:
                cnpjs_processados.add(cnpj)
            
            sigla = get_col_val(row, 'SIGLA')
            porte = get_col_val(row, 'PORTE')
            er = get_col_val(row, 'ER')
            carteira = get_col_val(row, 'CARTEIRA')
            endereco = get_col_val(row, 'ENDEREÇO', 'ENDERECO', 'LOGRADOURO')
            bairro = get_col_val(row, 'BAIRRO')
            municipio = get_col_val(row, 'MUNICIPIO', 'CIDADE')
            estado = get_col_val(row, 'ESTADO', 'UF')
            pais = get_col_val(row, 'PAIS', 'PAÍS')
            area = get_col_val(row, 'AREA', 'ÁREA')
            cnae_principal = get_col_val(row, 'CNAE PRINCIPAL', 'CNAE')
            descricao_cnae = get_col_val(row, 'DESCRIÇÃO CNAE', 'DESCRICAO CNAE')
            tipo_empresa = get_col_val(row, 'TIPO EMPRESA', 'TIPO')
            cep = get_col_val(row, 'CEP')
            
            try:
                num_func_val = get_col_val(row, 'NÚMERO FUNCIONÁRIOS', 'NUMERO FUNCIONARIOS', 'FUNCIONARIOS')
                numero_funcionarios = int(float(num_func_val)) if num_func_val else None
            except (ValueError, TypeError):
                numero_funcionarios = None
                
            observacao = get_col_val(row, 'OBSERVAÇÃO', 'OBSERVACAO', 'OBS')
            
            empresa_existente = None
            if cnpj:
                empresa_existente = db.query(Empresa).filter(Empresa.cnpj == cnpj).first()
            else:
                empresa_existente = db.query(Empresa).filter(Empresa.empresa.ilike(empresa_nome)).first()
            
            if empresa_existente:
                empresas_ignoradas += 1
            else:
                nova_empresa = Empresa(
                    empresa=empresa_nome,
                    cnpj=cnpj,
                    sigla=sigla,
                    porte=porte,
                    er=er,
                    carteira=carteira,
                    endereco=endereco,
                    bairro=bairro,
                    municipio=municipio,
                    estado=estado,
                    pais=pais,
                    area=area,
                    cnae_principal=cnae_principal,
                    descricao_cnae=descricao_cnae,
                    tipo_empresa=tipo_empresa,
                    numero_funcionarios=numero_funcionarios,
                    observacao=observacao,
                    cep=cep
                )
                db.add(nova_empresa)
                empresas_criadas += 1
        
        db.commit()
        
        return {
            "message": "Upload concluído com sucesso",
            "empresas_criadas": empresas_criadas,
            "empresas_ignoradas": empresas_ignoradas,
            "total_processadas": empresas_criadas + empresas_ignoradas
        }
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erro ao processar arquivo: {str(e)}"
        )
@router.get("/{empresa_id}/historico")
def obter_historico_empresa(
    empresa_id: int,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual)
):
    from backend.models.historico import HistoricoEmpresa
    from backend.models.usuarios import Usuario
    
    historico = db.query(HistoricoEmpresa).join(Usuario).filter(
        HistoricoEmpresa.empresa_id == empresa_id
    ).order_by(HistoricoEmpresa.data.desc()).all()
    
    return [
        {
            "id": h.id,
            "data": h.data,
            "tipo_acao": h.tipo_acao,
            "detalhes": h.detalhes,
            "usuario_nome": h.usuario.nome
        } for h in historico
    ]

@router.get("/{empresa_id}/programas")
def obter_programas_empresa(
    empresa_id: int,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual)
):
    from backend.models import CronogramaEvento, Program
    from sqlalchemy import func
    
    # Busca eventos da empresa agrupados por programa
    eventos = db.query(
        Program.nome,
        CronogramaEvento.program_id,
        func.count(CronogramaEvento.id).label("total"),
        func.min(CronogramaEvento.data).label("primeira_sessao"),
        func.max(CronogramaEvento.data).label("ultima_sessao")
    ).join(Program).filter(
        CronogramaEvento.empresa_id == empresa_id
    ).group_by(CronogramaEvento.program_id, Program.nome).all()
    
    resultados = []
    hoje = datetime.now().strftime("%Y-%m-%d")
    
    for ev in eventos:
        # Calcular realizados e próxima data
        total = ev.total
        
        realizados = db.query(func.count(CronogramaEvento.id)).filter(
            CronogramaEvento.empresa_id == empresa_id,
            CronogramaEvento.program_id == ev.program_id,
            CronogramaEvento.data <= hoje
        ).scalar()
        
        proxima = db.query(CronogramaEvento.data).filter(
            CronogramaEvento.empresa_id == empresa_id,
            CronogramaEvento.program_id == ev.program_id,
            CronogramaEvento.data > hoje
        ).order_by(CronogramaEvento.data).first()
        
        resultados.append({
            "nome": ev.nome,
            "total_sessoes": total,
            "sessoes_realizadas": realizados,
            "progresso": round((realizados / total) * 100) if total > 0 else 0,
            "proxima_sessao": proxima[0] if proxima else None,
            "status": "Concluído" if realizados == total else "Em Andamento"
        })
        
    return resultados
