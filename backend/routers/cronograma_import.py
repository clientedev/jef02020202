from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
from datetime import date, timedelta, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from backend.database import get_db
from backend.models import Empresa, Usuario
from backend.models.cronograma import Program, CronogramaEvento, CategoriaEvento, PeriodoEvento, CronogramaProjeto, CronogramaAtividade
from backend.auth.security import obter_usuario_atual

router = APIRouter(prefix="/api/cronograma", tags=["cronograma-import"])

COLUNAS_MODELO = [
    "TIPO DE PROGRAMA", "CNPJ", "EMPRESA", "PORTE", "ER", "SIGLA",
    "ETAPA", "Nº PROPOSTA", "SOLUÇÃO", "CH", "CONSULTOR",
    "DATA INÍCIO", "DATA TÉRMINO"
]

def _col(row_data: dict, *keys):
    """Get first matching key (case-insensitive)."""
    for k in keys:
        for rk, rv in row_data.items():
            if rk.strip().upper() == k.strip().upper():
                if rv is None:
                    continue
                # If it's already a date/datetime, return it as is
                if isinstance(rv, (date, datetime)):
                    return rv
                s = str(rv).strip()
                if s.lower() in ('none', 'nan', ''):
                    continue
                return s
    return None


@router.get("/import-template")
def download_template():
    """Serve the Excel import template for download."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cronograma"

    header_fill = PatternFill("solid", fgColor="1F7A1F")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, col_name in enumerate(COLUNAS_MODELO, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Example row
    example = [
        "SEG+", "12.345.678/0001-99", "Empresa Exemplo Ltda", "Pequeno", "SE",
        "EMP", "Diagnóstico", "PROP-001", "Lean Manufacturing", "40",
        "João Silva", "2025-04-01", "2025-06-30"
    ]
    for col_idx, val in enumerate(example, start=1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Column widths
    widths = [18, 20, 30, 10, 8, 10, 14, 14, 20, 8, 18, 13, 13]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    # Add note sheet
    ws_info = wb.create_sheet("Instruções")
    ws_info.column_dimensions['A'].width = 80
    instrucoes = [
        ("INSTRUÇÕES DE PREENCHIMENTO", True),
        ("", False),
        ("TIPO DE PROGRAMA: Nome do programa (ex: SEG+, LEAN, Diagnóstico...)", False),
        ("CNPJ: CNPJ da empresa (opcional, usado para identificar empresa já cadastrada)", False),
        ("EMPRESA: Nome da empresa (obrigatório)", False),
        ("PORTE: Pequeno, Médio, Grande, etc.", False),
        ("ER: Regional (ex: SE, SP, RJ...)", False),
        ("SIGLA: Sigla da empresa", False),
        ("ETAPA: Etapa do programa", False),
        ("Nº PROPOSTA: Número da proposta", False),
        ("SOLUÇÃO: Descrição da solução/programa", False),
        ("CH: Carga horária total em horas (ex: 40)", False),
        ("CONSULTOR: Nome completo do consultor (deve estar cadastrado no sistema)", False),
        ("DATA INÍCIO: Data de início no formato AAAA-MM-DD ou DD/MM/AAAA", False),
        ("DATA TÉRMINO: Data de término no formato AAAA-MM-DD ou DD/MM/AAAA", False),
        ("", False),
        ("IMPORTANTE: O dia da semana será calculado automaticamente a partir da DATA INÍCIO.", False),
        ("Os eventos serão lançados nesse mesmo dia da semana ao longo do período.", False),
        ("Se o programa não existir, será criado automaticamente.", False),
        ("Se a empresa não estiver cadastrada (pelo CNPJ ou nome), será criada automaticamente.", False),
    ]
    for i, (text, bold) in enumerate(instrucoes, start=1):
        cell = ws_info.cell(row=i, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, color="1F7A1F", size=12)
        else:
            cell.font = Font(size=10)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=modelo_cronograma.xlsx"}
    )


def _parse_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, (date, datetime)):
        if isinstance(val, datetime):
            return val.date()
        return val
    
    s = str(val).strip()
    # Handle 'YYYY-MM-DD HH:MM:SS' or similar formats by taking only the date part
    if ' ' in s:
        s = s.split(' ')[0]
    
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


@router.post("/import-excel")
async def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(obter_usuario_atual)
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Arquivo deve ser .xlsx ou .xls")

    contents = await file.read()
    wb = openpyxl.load_workbook(BytesIO(contents), data_only=True)
    ws = wb.active

    # Build header map  col_name -> col_index (0-based)
    headers = []
    for cell in ws[1]:
        v = str(cell.value).strip().upper() if cell.value else ""
        headers.append(v)

    results = {
        "rows_processed": 0,
        "eventos_criados": 0,
        "programas_criados": [],
        "empresas_criadas": [],
        "erros": []
    }

    # Cache consultores
    consultores_db = {c.nome.strip().upper(): c for c in db.query(Usuario).all()}

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None for v in row):
            continue

        row_data = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}

        tipo_programa = _col(row_data, "TIPO DE PROGRAMA", "PROGRAMA", "TIPO")
        cnpj_raw = _col(row_data, "CNPJ")
        empresa_nome = _col(row_data, "EMPRESA")
        porte = _col(row_data, "PORTE")
        er = _col(row_data, "ER")
        sigla = _col(row_data, "SIGLA")
        ch_raw = _col(row_data, "CH", "CARGA HORÁRIA", "CARGA HORARIA")
        consultor_nome = _col(row_data, "CONSULTOR")
        data_inicio_raw = _col(row_data, "DATA INÍCIO", "DATA INICIO", "DATA DE INÍCIO")
        data_termino_raw = _col(row_data, "DATA TÉRMINO", "DATA TERMINO", "DATA DE TÉRMINO")
        solucao = _col(row_data, "SOLUÇÃO", "SOLUCAO", "SOLUÇÃO/PROGRAMA")
        num_proposta = _col(row_data, "Nº PROPOSTA", "PROPOSTA", "NºP", "NUMERO PROPOSTA")

        if not empresa_nome:
            results["erros"].append(f"Linha {row_num}: EMPRESA é obrigatória")
            continue
        if not tipo_programa:
            results["erros"].append(f"Linha {row_num}: TIPO DE PROGRAMA é obrigatório")
            continue

        # Parse dates
        data_inicio = _parse_date(data_inicio_raw)
        data_termino = _parse_date(data_termino_raw)

        if not data_inicio:
            results["erros"].append(f"Linha {row_num}: DATA INÍCIO inválida: '{data_inicio_raw}'")
            continue
        if not data_termino:
            results["erros"].append(f"Linha {row_num}: DATA TÉRMINO inválida: '{data_termino_raw}'")
            continue

        # Parse CH
        try:
            carga_horaria = float(str(ch_raw).replace(",", ".")) if ch_raw else 40.0
        except ValueError:
            carga_horaria = 40.0

        # --- Find or create empresa ---
        empresa = None
        if cnpj_raw:
            cnpj_clean = ''.join(filter(str.isdigit, cnpj_raw))
            if cnpj_clean:
                empresa = db.query(Empresa).filter(Empresa.cnpj.ilike(f"%{cnpj_clean}%")).first()

        if not empresa:
            empresa = db.query(Empresa).filter(
                Empresa.empresa.ilike(f"%{empresa_nome.strip()}%")
            ).first()

        if not empresa:
            empresa = Empresa(
                empresa=empresa_nome,
                cnpj=cnpj_raw,
                porte=porte,
                er=er,
                sigla=sigla,
            )
            db.add(empresa)
            db.flush()  # get ID
            results["empresas_criadas"].append(empresa_nome)
        else:
            # Update missing fields
            if porte and not empresa.porte:
                empresa.porte = porte
            if er and not empresa.er:
                empresa.er = er
            if sigla and not empresa.sigla:
                empresa.sigla = sigla

        # --- Find or create program ---
        program = db.query(Program).filter(
            Program.nome.ilike(f"%{tipo_programa.strip()}%")
        ).first()

        if not program:
            program = Program(
                nome=tipo_programa,
                carga_horaria=carga_horaria,
                descricao=solucao or tipo_programa,
                numero_proposta=num_proposta or None,
            )
            db.add(program)
            db.flush()
            results["programas_criados"].append(tipo_programa)
        else:
            # Update carga_horaria if empty, and proposta if not set
            if program.carga_horaria == 0:
                program.carga_horaria = carga_horaria
            if num_proposta and not program.numero_proposta:
                program.numero_proposta = num_proposta

        # --- Find consultor ---
        consultor = None
        if consultor_nome:
            c_nome_up = consultor_nome.strip().upper()
            # 1. Try Exact Match
            consultor = consultores_db.get(c_nome_up)
            if not consultor:
                # 2. Try partial match but avoid being too greedy
                # Only match if the names share significant parts
                for nome_key, c in consultores_db.items():
                    # If Excel name is a word inside DB name OR vice-versa
                    parts = set(c_nome_up.split())
                    db_parts = set(nome_key.split())
                    if parts & db_parts: # At least one common word
                        consultor = c
                        break

        if not consultor:
            results["erros"].append(f"Linha {row_num}: Consultor '{consultor_nome}' não encontrado no sistema.")
            continue

        # --- Sync with CronogramaProjeto ---
        projeto = db.query(CronogramaProjeto).filter(
            CronogramaProjeto.empresa_id == empresa.id,
            CronogramaProjeto.consultor_id == consultor.id,
            CronogramaProjeto.proposta == (_col(row_data, "Nº PROPOSTA") or "Excel Import")
        ).first()

        if not projeto:
            projeto = CronogramaProjeto(
                proposta=_col(row_data, "Nº PROPOSTA") or "Excel Import",
                empresa_id=empresa.id,
                sigla=sigla or empresa.sigla,
                solucao=solucao or tipo_programa,
                horas_totais=carga_horaria,
                consultor_id=consultor.id,
                consultor_nome=consultor.nome,
                data_inicio=data_inicio,
                data_termino=data_termino,
            )
            db.add(projeto)
            db.flush()

        # --- Generate events based on date range and weekday ---
        # Calculate number of occurrences (weeks) based on duration
        diff_days = (data_termino - data_inicio).days
        num_semanas = (diff_days // 7) + 1
        
        # Calculate hours per session
        if num_semanas > 0:
            horas_por_sessao = round(carga_horaria / num_semanas, 2)
        else:
            horas_por_sessao = carga_horaria

        horas_restantes = carga_horaria
        eventos_linha = 0

        # Cache existing dates for this consultor in this import to avoid self-collisions
        dias_ocupados = set()

        for w in range(num_semanas + 2): # safety buffer
            if horas_restantes <= 0.05:
                break
            
            d_ideal = data_inicio + timedelta(weeks=w)
            d_evento = d_ideal

            # Collision Avoidance: Find next available business day
            intentos = 0
            while intentos < 30:
                # 1. Skip weekends
                if d_evento.weekday() >= 5:
                    d_evento += timedelta(days=1)
                    continue
                
                # 2. Check DB for existing events (one company per day rule)
                conflito_db = db.query(CronogramaEvento).filter(
                    CronogramaEvento.consultor_id == consultor.id,
                    CronogramaEvento.data == d_evento,
                    CronogramaEvento.empresa_id != empresa.id # allowed if same company (rare)
                ).first()

                if conflito_db or d_evento in dias_ocupados:
                    d_evento += timedelta(days=1)
                    intentos += 1
                    continue
                
                break # Found a free day

            if d_evento > data_termino + timedelta(weeks=4): # Don't drift too far
                break

            dias_ocupados.add(d_evento)

            h_hoje = min(horas_por_sessao, horas_restantes)
            if h_hoje <= 0: break

            periodo = PeriodoEvento.dia_todo if h_hoje >= 6 else PeriodoEvento.manha

            evento = CronogramaEvento(
                data=d_evento,
                categoria=CategoriaEvento.programado,
                periodo=periodo,
                empresa_id=empresa.id,
                sigla_empresa=empresa.sigla or sigla or "",
                consultor_id=consultor.id,
                program_id=program.id,
                projeto_id=projeto.id,
                titulo=f"{tipo_programa} - {empresa_nome}",
                descricao=solucao or f"Sessão de {tipo_programa}",
                carga_horaria=h_hoje,
            )
            db.add(evento)
            horas_restantes -= h_hoje
            eventos_linha += 1

        results["eventos_criados"] += eventos_linha
        results["rows_processed"] += 1

    db.commit()
    return results
